"""Excel: каждый лист разворачивается в текстовую таблицу с заголовками."""
from __future__ import annotations

import io

MAX_ROWS = 200
MAX_COLS = 30


def _fmt(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def extract(data: bytes) -> tuple[str, list[str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    notes: list[str] = []
    chunks: list[str] = []

    for sheet in workbook.worksheets:
        rows_out: list[str] = []
        for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_index >= MAX_ROWS:
                notes.append(f"Лист «{sheet.title}» обрезан до {MAX_ROWS} строк")
                break
            cells = [_fmt(v) for v in row[:MAX_COLS]]
            if not any(cells):
                continue
            rows_out.append(" | ".join(cells).rstrip(" |"))
        if rows_out:
            chunks.append(f"Лист: {sheet.title}")
            chunks.extend(rows_out)

    workbook.close()
    return "\n".join(chunks), notes
